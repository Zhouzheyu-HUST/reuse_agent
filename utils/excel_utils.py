# Copyright (c) Huawei Technologies Co., Ltd. 2025. All rights reserved.

__author__ = "Zhipeng Hou"

import os
from typing import Union

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

from utils import (
    read_json,
    to_float,
    to_bool,
    mean_metric,
    ratio_true
)


class ExcelOperation(object):

    sheet_head_path = "configs/eval_excel_config.json"
    sheet_head = read_json(sheet_head_path)

    def __init__(
        self,
        sheet_name: str, 
        save_path: str,
        start_row: int
    ) -> None:
        self.save_path = save_path
        self.start_row = start_row
        self.flat_columns = self._get_head_data()
        self.st = self._get_styles()
        self.wb, self.ws = self._get_workbook(sheet_name, save_path)
    
    @staticmethod
    def _get_styles() -> dict:
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        vcenter_wrap = Alignment(vertical="center", wrap_text=True)

        fill_group = PatternFill("solid", fgColor="BFBFBF")
        fill_head = PatternFill("solid", fgColor="D9D9D9")

        thin = Side(style="thin", color="808080")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        return {
            "bold": bold,
            "center": center,
            "vcenter_wrap": vcenter_wrap,
            "fill_group": fill_group,
            "fill_head": fill_head,
            "border": border,
        }

    @staticmethod
    def _compute_metrics(rows: list[dict[str, any]]) -> dict:
        pass_1_cache_ratio = []
        pass_3_cache_ratio = []

        pass_1_step_ratio = []
        pass_3_step_ratio = []

        pass_1_single_step_time = []
        pass_3_single_step_time = []

        pass_1_total_step_time = []
        pass_3_total_step_time = []

        pass_1_acc = []
        pass_3_acc = []

        pass_1_cache_exec_ratio = []
        pass_3_cache_exec_ratio = []

        for d in rows:
            cur_pass_1_cache_ratio = to_bool(d.get("pass@1 是否使用缓存"))
            pass_1_cache_ratio.append(cur_pass_1_cache_ratio)

            cur_pass_3_cache_ratio = to_bool(d.get("pass@3 是否使用缓存"))
            pass_3_cache_ratio.append(cur_pass_3_cache_ratio)

            cur_pass_1_step_ratio = to_float(d.get("pass@1 操控步骤数")) / to_float(d.get("操控步数(GT)"))
            pass_1_step_ratio.append(cur_pass_1_step_ratio)

            cur_pass_3_step_ratio = to_float(d.get("pass@3 操控步骤数")) / to_float(d.get("操控步数(GT)"))
            pass_3_step_ratio.append(cur_pass_3_step_ratio)

            cur_pass_1_single_step_time = to_float(d.get("pass@1 单步操控时延(s)"))
            pass_1_single_step_time.append(cur_pass_1_single_step_time)

            cur_pass_3_single_step_time = to_float(d.get("pass@3 单步操控时延(s)"))
            pass_3_single_step_time.append(cur_pass_3_single_step_time)

            cur_pass_1_total_step_time = to_float(d.get("pass@1 端到端耗时(s)"))
            pass_1_total_step_time.append(cur_pass_1_total_step_time)

            cur_pass_3_total_step_time = to_float(d.get("pass@3 端到端耗时(s)"))
            pass_3_total_step_time.append(cur_pass_3_total_step_time)

            cur_pass_1_acc = to_bool(d.get("pass@1 是否成功"))
            pass_1_acc.append(cur_pass_1_acc)

            cur_pass_3_acc = to_bool(d.get("pass@3 是否成功"))
            pass_3_acc.append(cur_pass_3_acc)

            pass_1_cache_exec_ratio.append(cur_pass_1_acc and cur_pass_1_cache_ratio)
            pass_3_cache_exec_ratio.append(cur_pass_3_acc and cur_pass_3_cache_ratio)

        
        metrics = {
            "pass@1 任务缓存命中率": round(ratio_true(pass_1_cache_ratio), 1),
            "pass@3 任务缓存命中率": round(ratio_true(pass_3_cache_ratio), 1),
            "pass@1 任务缓存执行成功率": round(ratio_true(pass_1_cache_exec_ratio), 1),
            "pass@3 任务缓存执行成功率": round(ratio_true(pass_3_cache_exec_ratio), 1),
            "pass@1 智能体与人操控步数比": round(mean_metric(pass_1_step_ratio), 1),
            "pass@3 智能体与人操控步数比": round(mean_metric(pass_3_step_ratio), 1),
            "pass@1 单步操控时延(s)": round(mean_metric(pass_1_single_step_time), 4),
            "pass@3 单步操控时延(s)": round(mean_metric(pass_3_single_step_time), 4),
            "pass@1 端到端耗时(s)": round(mean_metric(pass_1_total_step_time), 4),
            "pass@3 端到端耗时(s)": round(mean_metric(pass_3_total_step_time), 4),
            "pass@1 任务成功率": round(ratio_true(pass_1_acc), 2),
            "pass@3 任务成功率": round(ratio_true(pass_3_acc), 2)
        }
        return metrics

    def _build_col_idx(self) -> dict[str, int]:
        idx: dict[str, int] = {}
        max_col = self.ws.max_column
        for c in range(1, max_col + 1):
            v = self.ws.cell(row=self.start_row - 1, column=c).value
            if v is None:
                continue
            name = str(v).strip()
            if name in self.flat_columns:
                idx[name] = c

        missing = [c for c in self.flat_columns if c not in idx]
        if missing:
            raise ValueError(f"Excel表头缺失这些列：{missing}")
        return idx

    def _iter_rows_as_dicts(self) -> list[dict[str, any]]:
        col_index = self._build_col_idx()

        rows: list[dict[str, any]] = []
        r = self.start_row
        while True:
            row_dict: dict[str, any] = {}
            empty = True
            for col_name in self.flat_columns:
                c = col_index[col_name]
                v = self.ws.cell(row=r, column=c).value
                row_dict[col_name] = v
                if v not in (None, ""):
                    empty = False

            if empty:
                break
            rows.append(row_dict)
            r += 1
        return rows
    
    def _get_head_data(self) -> list[str]:
        flat_columns: list[str] = []
        for _, cols in self.sheet_head.items():
            flat_columns.extend(cols)
        
        return flat_columns

    def _create_sheet_heads(
        self,
        ws: Worksheet
    ) -> None:
        if ws["A2"].value is not None:
            return

        # 行高
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 30

        col_idx = 1
        for group_name, cols in self.sheet_head.items():
            start_col = col_idx
            end_col = col_idx + len(cols) - 1

            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            cell = ws.cell(row=1, column=start_col)
            cell.value = group_name
            cell.font = self.st["bold"]
            cell.alignment = self.st["center"]
            cell.fill = self.st["fill_group"]

            col_idx = end_col + 1

        total_cols = len(self.flat_columns)
        for i, col_name in enumerate(self.flat_columns, start=1):
            cell = ws.cell(row=2, column=i)
            cell.value = col_name
            cell.font = self.st["bold"]
            cell.alignment = self.st["center"]
            cell.fill = self.st["fill_head"]
        
        for r in range(1, 3):
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).border = self.st["border"]
        
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=total_cols).coordinate}"

        for c in range(1, total_cols + 1):
            title = ws.cell(row=2, column=c).value or ""
            w = max(10, min(30, len(str(title)) * 2))
            ws.column_dimensions[ws.cell(row=2, column=c).column_letter].width = w

    def _get_workbook(
        self,
        sheet_name: str, 
        save_path: str
    ) -> tuple[Workbook, Worksheet]:
        if os.path.exists(save_path):
            wb = load_workbook(save_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        self._create_sheet_heads(ws)
        
        return wb, ws

    def _row_is_empty(
        self,
        row_idx: int
    ) -> bool:
        for c in range(1, len(self.flat_columns) + 1):
            v = self.ws.cell(row=row_idx, column=c).value
            if v not in (None, ""):
                return False
        return True

    def _next_data_row(self) -> int:
        r = self.start_row
        while not self._row_is_empty(r):
            r += 1
        return r

    def save(self) -> None:
        self.wb.save(self.save_path)

    def insert_one_raw(
        self,
        row_values: Union[list, tuple, dict]
    ) -> None:
        r = self._next_data_row()

        if isinstance(row_values, dict):
            # dict：按列名对齐
            values = [row_values.get(col_name, "") for col_name in self.flat_columns]
        else:
            # list/tuple：不足补空，超出截断
            values = list(row_values)[:len(self.flat_columns)] + [""] * max(0, len(self.flat_columns) - len(row_values))
        
        for col_idx, v in enumerate(values, start=1):
            cell = self.ws.cell(row=r, column=col_idx)
            cell.value = v
            cell.border = self.st["border"]
            cell.alignment = self.st["center"]
    
    def write_overall_excel(
        self,
        sheet_name: str, 
        save_path: str
    ) -> None:
        rows = self._iter_rows_as_dicts()
        metrics = self._compute_metrics(rows)

        dir_ = os.path.dirname(save_path)
        if dir_:
            os.makedirs(dir_, exist_ok=True)
        
        mode = "a" if os.path.exists(save_path) else "w"

        writer_kwargs: dict[str, any] = {"engine": "openpyxl", "mode": mode}
        if mode == "a":
            # 只有 append 模式才允许
            writer_kwargs["if_sheet_exists"] = "replace"

        with pd.ExcelWriter(save_path, **writer_kwargs) as w:
            pd.DataFrame([metrics]).to_excel(w, sheet_name=sheet_name, index=False)
